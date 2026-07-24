"""Edge fetcher for spf (Session 8) — Philadelphia Fed Survey of Professional Forecasters.

The SPF median inflation forecasts (CPI, core CPI, PCE, core PCE) are QUARTERLY, annualized-rate
forecasts at horizons 0 (current quarter) .. 5 (five quarters ahead), keyed by the survey's
(year, quarter). Role: a quarterly TRAJECTORY benchmark, NOT a per-print instrument — it is
reported in its own table and never aligned to a monthly MoM surprise.

Public domain (U.S. Fed). Derived eval artifact -> data/benchmarks/spf.csv (like cleveland_nowcast).
"""
from __future__ import annotations

import csv
import datetime as dt
import io
import sys
import zipfile
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
import _ingest  # noqa: E402

PIPE = Path(__file__).resolve().parent
REPO = Path(__file__).resolve().parents[2]
URL = ("https://www.philadelphiafed.org/-/media/FRBP/Assets/Surveys-And-Data/"
       "survey-of-professional-forecasters/historical-data/medianLevel.xlsx")
OUT_CSV = REPO / "data" / "benchmarks" / "spf.csv"
VARS = {"CPI": "cpi", "CORECPI": "core_cpi", "PCE": "pce", "COREPCE": "core_pce"}
COLS = ["source", "variable", "survey_year", "survey_quarter", "horizon_q",
        "value_annualized_pct", "vintage_status", "observed_asof"]

# SPF survey deadlines land mid-2nd-month of the quarter; a conservative point-in-time as-of.
_SURVEY_MONTH = {1: 2, 2: 5, 3: 8, 4: 11}


def _strip_core_props(raw: bytes) -> bytes:
    """openpyxl rejects the workbook's date-formatted docProps/core.xml; drop it (harmless)."""
    zin = zipfile.ZipFile(io.BytesIO(raw))
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zout:
        for it in zin.infolist():
            if it.filename != "docProps/core.xml":
                zout.writestr(it, zin.read(it.filename))
    return buf.getvalue()


def parse(raw: bytes) -> list[dict]:
    """medianLevel.xlsx -> one row per (variable, survey, horizon). Pure/testable."""
    wb = openpyxl.load_workbook(io.BytesIO(_strip_core_props(raw)), read_only=True, data_only=True)
    rows = []
    for sheet, var in VARS.items():
        if sheet not in wb.sheetnames:
            continue
        it = wb[sheet].iter_rows(values_only=True)
        header = next(it)
        # horizon columns are {SHEET}1..{SHEET}6 = current quarter .. +5
        hcols = {i: int(h[len(sheet):]) - 1 for i, h in enumerate(header)
                 if isinstance(h, str) and h.startswith(sheet) and h[len(sheet):].isdigit()}
        for r in it:
            if not r or r[0] is None or r[1] is None:
                continue
            try:
                y, q = int(r[0]), int(r[1])
            except (TypeError, ValueError):
                continue
            asof = dt.date(y, _SURVEY_MONTH[q], 15).isoformat()
            for i, hz in hcols.items():
                v = r[i] if i < len(r) else None
                if isinstance(v, (int, float)):
                    rows.append({"source": "spf", "variable": var, "survey_year": y,
                                 "survey_quarter": q, "horizon_q": hz,
                                 "value_annualized_pct": f"{float(v):.4f}",
                                 "vintage_status": "point_in_time", "observed_asof": asof})
    wb.close()
    return rows


def write_csv(rows: list[dict], path: Path = OUT_CSV) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=COLS)
        w.writeheader()
        w.writerows(rows)


def fetch(as_of: str | None = None):
    as_of = as_of or dt.date.today().isoformat()
    out = _ingest.raw_dir("spf", as_of)
    raw, status = _ingest.fetch(URL)
    (out / "medianLevel.xlsx").write_bytes(raw)
    prov = [{"label": "spf_median_level", "source_url": URL, "http_status": status,
             "retrieved_at_utc": dt.datetime.now(dt.timezone.utc).isoformat(),
             "sha256": _ingest.sha256(raw), "bytes": len(raw)}]
    _ingest.write_provenance(out, prov)
    rows = parse(raw)
    write_csv(rows)
    print(f"spf: {len(rows)} (variable,survey,horizon) rows -> {OUT_CSV.relative_to(REPO)}")
    return rows


if __name__ == "__main__":
    fetch()
