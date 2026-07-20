"""Edge fetcher for manheim (Session 2B-final, B1) — POINT-IN-TIME UVVI archive.

Each dated monthly xlsx is the index as published that month; its newest reference-month
row is that month's FIRST-RELEASE full-month value. We enumerate dated publications,
keep those still hosted (older 404 -> honest short coverage), and emit one point_in_time
row per publication. Full-month only; mid-month is a separate publication (H1 follow-up).
"""
from __future__ import annotations
import datetime as dt
import io
import sys
from pathlib import Path

import yaml

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
import _ingest  # noqa: E402

PIPE = Path(__file__).resolve().parent
_MON = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def _pub_months(start_ym: str) -> list[tuple[int, int]]:
    y, m = (int(x) for x in start_ym.split("-"))
    today = dt.date.today()
    out = []
    while (y, m) <= (today.year, today.month):
        out.append((y, m))
        m += 1
        if m > 12:
            y += 1
            m = 1
    return out


def full_series(xlsx_bytes: bytes, spec: dict) -> list[tuple[str, float]]:
    """ALL (reference_period first-of-month, index) rows from a UVVI xlsx DATA sheet.

    Used for the historical ingest (Session-3A Task 0b): the point-in-time archive's 11
    months matched the latest download EXACTLY (see license_note), so the full-month MUVVI
    is unrevised post-publication and the historical download equals the first-release
    series. Pure/testable."""
    import openpyxl

    a = spec["archive"]
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    ws = wb[a["data_sheet"]]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d, val = row[a["date_col_index"]], row[a["index_col_index"]]
        if isinstance(d, dt.datetime) and isinstance(val, (int, float)):
            out.append((d.date().replace(day=1).isoformat(), float(val)))
    return sorted(out)


def extract_newest(xlsx_bytes: bytes, spec: dict) -> tuple[str, float] | None:
    """Newest (reference_period first-of-month, index value) from a dated UVVI xlsx. Pure."""
    import openpyxl

    a = spec["archive"]
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    ws = wb[a["data_sheet"]]
    best = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        d, val = row[a["date_col_index"]], row[a["index_col_index"]]
        if isinstance(d, dt.datetime) and isinstance(val, (int, float)):
            period = d.date().replace(day=1)
            if best is None or period > best[0]:
                best = (period, float(val))
    return (best[0].isoformat(), best[1]) if best else None


def fetch(as_of: str | None = None):
    spec = yaml.safe_load((PIPE / "spec.yaml").read_text())
    as_of = as_of or dt.date.today().isoformat()
    out = _ingest.raw_dir(spec["source"], as_of)
    a = spec["archive"]
    rows, prov, quarantine, seen = [], [], 0, set()
    for py, pm in _pub_months(a["enumerate_pub_from"]):
        rm, ry = (pm - 1, py) if pm > 1 else (12, py - 1)  # ref month = pub month - 1
        url = a["url_template"].format(pub_year=py, pub_month=pm, ref_mon=_MON[rm - 1], ref_year=ry)
        try:
            raw, status = _ingest.fetch(url, timeout=60)
        except Exception:
            continue  # 404 / not hosted -> honest coverage gap, not interpolated
        (out / f"{_MON[rm - 1]}-{ry}-UVVI.xlsx").write_bytes(raw)
        try:
            rec = extract_newest(raw, spec)
        except Exception:
            quarantine += 1
            continue  # parse failure quarantined, never partial rows
        if rec is None or rec[0] in seen:
            continue
        seen.add(rec[0])
        period, value = rec
        rows.append({"source": spec["source"], "series_key": spec["series_key"],
                     "frequency": spec["frequency"], "period": period, "value": f"{value:.6f}",
                     "vintage_status": spec["vintage_status"],
                     "observed_date": dt.date(py, pm, 1).isoformat()})
        prov.append({"label": f"UVVI_{period}", "source_url": url, "http_status": status,
                     "retrieved_at_utc": dt.datetime.now(dt.timezone.utc).isoformat(),
                     "sha256": _ingest.sha256(raw), "bytes": len(raw)})
    prov_path = _ingest.write_provenance(out, prov)
    rows.sort(key=lambda r: r["period"])
    staged = out / "staged.csv"
    _ingest.write_staged_csv(staged, rows)
    if rows:
        print(f"manheim: {len(rows)} point-in-time months {rows[0]['period']}..{rows[-1]['period']} "
              f"| quarantined {quarantine}")
    else:
        print("manheim: 0 rows")
    n = _ingest.load(spec["source"], staged, prov_path, as_of)
    print(f"manheim: loaded {n} rows")
    return staged


if __name__ == "__main__":
    fetch(sys.argv[1] if len(sys.argv) > 1 else None)
