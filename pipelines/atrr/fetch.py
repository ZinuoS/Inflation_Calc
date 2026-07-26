"""Edge fetcher for atrr (Task 3 follow-on) — BLS research CPI for all-tenant regressed rent.

Two jobs:
  1. **Recurring snapshot.** Pull the current workbook and archive an immutable full-history vintage
     (`data/raw/atrr/vintage_{date}/`). The source is *"perpetually revised"* (BLS's own words), so a
     restated history cannot be backtested from the latest file alone — the snapshot IS the evidence.
     Byte-identical payloads are skipped: publication is paused, so "no new vintage" is the honest
     signal, not a reason to store duplicates.
  2. **Backfill BLS's own archive.** BLS publishes dated per-quarter files
     (`...-{YYYY}q{N}.xlsx`, 2023q3 →). Those are genuine as-published vintages (verified: 1999q4
     reads 102.2753677 in the 2024q2 file vs 102.388 in the current one). Archiving them gives a real
     vintage history immediately, instead of waiting four quarters for our own to accrue.

Parsed output: quarterly R-CPI-ATR **index levels** → `proxy_observations` (period = first month of the
quarter). The published 4-quarter-change columns are deliberately NOT used — hard rule 8 bars YoY
quantities; quarterly changes are derived from levels at evaluation time.
"""
from __future__ import annotations

import datetime as dt
import io
import re
import sys
import zipfile
from pathlib import Path

import openpyxl
import yaml

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
import _ingest  # noqa: E402

PIPE = Path(__file__).resolve().parent
REPO = Path(__file__).resolve().parents[2]
SHEET = "R-CPI-ATR"
QMONTH = {"q1": "01", "q2": "04", "q3": "07", "q4": "10"}


def _clean_xlsx(raw: bytes) -> openpyxl.Workbook:
    """openpyxl rejects these workbooks' date-formatted docProps; drop that member (harmless)."""
    zin = zipfile.ZipFile(io.BytesIO(raw))
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zout:
        for it in zin.infolist():
            if it.filename != "docProps/core.xml":
                zout.writestr(it, zin.read(it.filename))
    return openpyxl.load_workbook(buf, read_only=True, data_only=True)


def parse(raw: bytes, spec: dict, observed_date: str) -> list[dict]:
    """Workbook -> staged rows (quarterly index levels). Pure/testable."""
    wb = _clean_xlsx(raw)
    if SHEET not in wb.sheetnames:
        wb.close()
        raise ValueError(f"{SHEET} sheet absent; sheets={wb.sheetnames}")
    rows = []
    for r in wb[SHEET].iter_rows(values_only=True):
        q, val = (r[0] if r else None), (r[1] if r and len(r) > 1 else None)
        if not isinstance(q, str):
            continue
        m = re.fullmatch(r"(\d{4})(q[1-4])", q.strip())
        if not m or not isinstance(val, (int, float)):
            continue
        rows.append({"source": spec["source"], "series_key": "R-CPI-ATR",
                     "frequency": "quarterly",
                     "period": f"{m.group(1)}-{QMONTH[m.group(2)]}-01",
                     "value": f"{float(val):.6f}",
                     "vintage_status": spec["vintage_status"],
                     "observed_date": observed_date})
    wb.close()
    return rows


def archived_quarters(spec: dict) -> list[str]:
    """Quarter tags BLS currently offers as dated archive files (scraped from the landing page)."""
    raw, _ = _ingest.fetch(spec["landing_page"])
    tags = set(re.findall(r"r-cpi-ntr-and-r-cpi-atr-(\d{4}q[1-4])\.xlsx",
                          raw.decode("utf-8", "replace")))
    return sorted(tags)


def _capture(source: str, tag: str, raw: bytes, url: str, spec: dict, kind: str) -> None:
    rows = parse(raw, spec, observed_date=tag)
    # An EMPTY ATR column is a real vintage fact, not a failure: in the 2023q3 file every R-CPI-ATR
    # cell is "-" (the series was not yet populated). Archive it anyway — "the series was empty as of
    # this vintage" is exactly the kind of evidence the archive exists to preserve.
    try:
        v = _ingest.archive_vintage(
            source, tag, raw, f"r-cpi-ntr-and-r-cpi-atr-{tag}.xlsx", url,
            rows=len(rows),
            period_min=rows[0]["period"] if rows else None,
            period_max=rows[-1]["period"] if rows else None,
            extra={"vintage_kind": kind, "sheet": SHEET,
                   "atr_empty": not rows})
        latest = rows[-1]["period"] if rows else "none (ATR column empty in this vintage)"
        print(f"  vintage captured: {v.name} ({len(rows)} quarters, latest {latest})")
    except _ingest.VintageExists as e:
        print(f"  already captured: {e}")
    except _ingest.VintageUnchanged as e:
        print(f"  unchanged, not re-archived: {e}")


def fetch(as_of: str | None = None, backfill: bool = True):
    spec = yaml.safe_load((PIPE / "spec.yaml").read_text())
    as_of = as_of or dt.date.today().isoformat()
    src = spec["source"]

    # 1. BLS's own dated archive (once; each is an as-published vintage)
    if backfill:
        for tag in archived_quarters(spec):
            url = spec["archive_url_template"].format(tag=tag)
            raw, _ = _ingest.fetch(url)
            _capture(src, tag, raw, url, spec, kind="bls_archived_quarter")

    # 2. the current file -> recurring snapshot + the load
    out = _ingest.raw_dir(src, as_of)
    raw, status = _ingest.fetch(spec["url"])
    (out / "r-cpi-ntr-and-r-cpi-atr.xlsx").write_bytes(raw)
    rows = parse(raw, spec, observed_date=as_of)
    prov = [{"label": "r_cpi_atr_current", "source_url": spec["url"], "http_status": status,
             "retrieved_at_utc": dt.datetime.now(dt.timezone.utc).isoformat(),
             "sha256": _ingest.sha256(raw), "bytes": len(raw)}]
    prov_path = _ingest.write_provenance(out, prov)
    _capture(src, as_of, raw, spec["url"], spec, kind="our_pull_date")

    staged = out / "staged.csv"
    _ingest.write_staged_csv(staged, rows)
    print(f"atrr: {len(rows)} quarterly rows {rows[0]['period']}..{rows[-1]['period']}")
    n = _ingest.load(src, staged, prov_path, as_of)
    print(f"atrr: loaded {n} rows into proxy_observations")
    return staged


if __name__ == "__main__":
    fetch()
